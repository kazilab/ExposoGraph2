"""Generate supplementary interaction-engine audit tables for manuscript results."""

from __future__ import annotations

import csv
import sys
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from ExposoGraph.interaction_engine import (  # noqa: E402
    EXPOSURE_PROFILES,
    compute_interaction_matrix,
    decompose_synergy,
)

OUT_DIR = ROOT / "Supplementary_Tables"
XLSX_PATH = OUT_DIR / "Supplementary_Table_S6_ExposoGraph_interactions.xlsx"
CSV_SCENARIOS = OUT_DIR / "Supplementary_Table_S6_interaction_scenarios.csv"
CSV_DECOMPOSITION = OUT_DIR / "Supplementary_Table_S6_synergy_decomposition.csv"


PROFILE_NAMES = [
    "smoker",
    "moderate_drinker",
    "smoker_moderate_drinker",
    "smoker_heavy_drinker",
    "industrial_worker",
    "smoker_industrial_worker",
    "JHBUI_10030",
]

SELECTED_PAIRS = [
    ("smoker_moderate_drinker", "NNK_x_acetaldehyde"),
    ("smoker_moderate_drinker", "PAH_x_NNK"),
    ("smoker_moderate_drinker", "PAH_x_cadmium"),
    ("smoker_moderate_drinker", "PAH_x_benzene"),
    ("smoker_moderate_drinker", "benzene_x_NDMA"),
    ("smoker_heavy_drinker", "NNK_x_benzene"),
    ("smoker_heavy_drinker", "PAH_x_NNK"),
    ("JHBUI_10030", "PAH_x_NNK"),
    ("JHBUI_10030", "PAH_x_cadmium"),
    ("smoker_industrial_worker", "PAH_x_acrolein"),
]

PAIR_AUDITS = [
    {
        "label": "PAH with dioxin/TCDD induction",
        "exposure": {"PAH": 3.0, "TCDD": 1.0},
        "lifestyle": {"TCDD_exposed": True},
        "expected_note": (
            "TCDD/dioxin is represented as induction, not as a baseline-risk "
            "carcinogen pair; no pairwise synergy entry is produced."
        ),
    },
    {
        "label": "Benzene plus trichloroethylene/TCE",
        "exposure": {"benzene": 10.0, "trichloroethylene": 5.0},
        "lifestyle": {},
        "expected_note": (
            "Trichloroethylene is parameterized as a CYP2E1 substrate in the JSON, "
            "but is not currently mapped into compute_interaction_matrix."
        ),
    },
    {
        "label": "Benzene plus vinyl chloride",
        "exposure": {"benzene": 10.0, "vinyl_chloride": 5.0},
        "lifestyle": {},
        "expected_note": (
            "Supported chlorinated-solvent pair in current interaction matrix; "
            "modeled as antagonistic via CYP2E1 competition."
        ),
    },
]


def _format_mapping(mapping: dict[str, Any]) -> str:
    return "; ".join(f"{key}={value}" for key, value in mapping.items())


def _scenario_rows() -> list[dict[str, Any]]:
    rows = []
    for profile_name in PROFILE_NAMES:
        cfg = EXPOSURE_PROFILES[profile_name]
        result = compute_interaction_matrix(
            cfg["exposure"],
            lifestyle=cfg.get("lifestyle", {}),
            genotypes=cfg.get("genotypes", {}),
            tissue="Liver",
        )
        top_pairs = sorted(
            result.synergy_matrix.items(),
            key=lambda item: item[1],
            reverse=True,
        )[:8]
        for rank, (pair, score) in enumerate(top_pairs, start=1):
            rows.append(
                {
                    "profile": profile_name,
                    "rank": rank,
                    "pair": pair,
                    "pair_synergy": score,
                    "overall_interaction_factor": result.interaction_factor,
                    "total_independent_risk": result.total_independent_risk,
                    "total_interaction_risk": result.total_interaction_risk,
                    "gsh_fraction_normal": result.gsh_status.fraction_normal,
                    "gsh_tipping_point": result.gsh_status.tipping_point_reached,
                    "active_inducers": "; ".join(result.induction_effects.active_inducers),
                    "enzyme_folds": _format_mapping(result.induction_effects.enzyme_folds),
                    "exposure_profile": _format_mapping(cfg["exposure"]),
                    "genotypes": _format_mapping(cfg.get("genotypes", {})),
                }
            )
    return rows


def _decomposition_rows() -> list[dict[str, Any]]:
    rows = []
    for profile_name, pair in SELECTED_PAIRS:
        cfg = EXPOSURE_PROFILES[profile_name]
        decomposed = decompose_synergy(
            cfg["exposure"],
            lifestyle=cfg.get("lifestyle", {}),
            genotypes=cfg.get("genotypes", {}),
            tissue="Liver",
        )
        if pair not in decomposed:
            continue
        dec = decomposed[pair]
        rows.append(_decomposition_row(profile_name, dec))

    for audit in PAIR_AUDITS:
        result = compute_interaction_matrix(
            audit["exposure"],
            lifestyle=audit["lifestyle"],
            tissue="Liver",
        )
        if result.synergy_matrix:
            for pair, score in result.synergy_matrix.items():
                rows.append(
                    {
                        "profile": audit["label"],
                        "pair": pair,
                        "composite": score,
                        "dominant_mechanism": "",
                        "main_effect_induction": "",
                        "main_effect_competition": "",
                        "main_effect_gsh": "",
                        "interaction_induction_competition": "",
                        "interaction_induction_gsh": "",
                        "interaction_competition_gsh": "",
                        "interaction_three_way": "",
                        "reconstruction_residual": "",
                        "shapley_residual": "",
                        "residual_policy": "",
                        "state_count": "",
                        "compatibility_delta_comp": "",
                        "compatibility_delta_gsh": "",
                        "compatibility_delta_ind": "",
                        "compatibility_policy": "",
                        "note": audit["expected_note"],
                    }
                )
        else:
            rows.append(
                {
                    "profile": audit["label"],
                    "pair": "none",
                    "composite": result.interaction_factor,
                    "dominant_mechanism": "",
                    "main_effect_induction": "",
                    "main_effect_competition": "",
                    "main_effect_gsh": "",
                    "interaction_induction_competition": "",
                    "interaction_induction_gsh": "",
                    "interaction_competition_gsh": "",
                    "interaction_three_way": "",
                    "reconstruction_residual": "",
                    "shapley_residual": "",
                    "residual_policy": "",
                    "state_count": "",
                    "compatibility_delta_comp": "",
                    "compatibility_delta_gsh": "",
                    "compatibility_delta_ind": "",
                    "compatibility_policy": "",
                    "note": audit["expected_note"],
                }
            )
    return rows


def _decomposition_row(profile: str, dec: Any) -> dict[str, Any]:
    return {
        "profile": profile,
        "pair": dec.pair,
        "composite": dec.composite,
        "dominant_mechanism": dec.dominant_mechanism,
        "main_effect_induction": dec.main_effects["induction"],
        "main_effect_competition": dec.main_effects["competition"],
        "main_effect_gsh": dec.main_effects["gsh"],
        "interaction_induction_competition": dec.pairwise_interactions["induction+competition"],
        "interaction_induction_gsh": dec.pairwise_interactions["induction+gsh"],
        "interaction_competition_gsh": dec.pairwise_interactions["competition+gsh"],
        "interaction_three_way": dec.three_way_interaction,
        "reconstruction_residual": dec.reconstruction_residual,
        "shapley_residual": dec.shapley_residual,
        "residual_policy": dec.residual_policy,
        "state_count": len(dec.state_values),
        "compatibility_delta_comp": dec.delta_comp,
        "compatibility_delta_gsh": dec.delta_gsh,
        "compatibility_delta_ind": dec.delta_ind,
        "compatibility_policy": dec.compatibility_fields["policy"],
    }


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return _format_mapping(value)
    return value


def _add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([_cell_value(row.get(header)) for header in headers])

    header_fill = PatternFill("solid", fgColor="4C3A20")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for column_cells in ws.columns:
        header = str(column_cells[0].value)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200]
        )
        width = min(max(max_len + 2, len(header) + 2, 10), 65)
        ws.column_dimensions[column_cells[0].column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName=f"T_{title[:24]}", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium7", showRowStripes=True)
    ws.add_table(table)


def _add_metadata_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Metadata_References")
    rows = [
        ("Generated", date.today().isoformat()),
        ("API", "ExposoGraph.interaction_engine.compute_interaction_matrix"),
        ("Decomposition API", "ExposoGraph.interaction_engine.decompose_synergy"),
        ("Parameter source", "ExposoGraph/data/interaction_parameters.json"),
        (
            "Formula",
            "pair_synergy = adjusted pair risk / independent pair risk; "
            "decomposition reports Shapley main effects, pairwise mechanism interactions, "
            "the three-way term, and numerical reconstruction residual.",
        ),
        (
            "Important caveat",
            "Dioxin/TCDD is currently modeled as CYP induction, not as a pairwise "
            "baseline-risk carcinogen in the interaction matrix.",
        ),
        (
            "Important caveat",
            "Trichloroethylene/TCE has CYP2E1 parameters in JSON but is not currently "
            "mapped into compute_interaction_matrix as a present carcinogen.",
        ),
    ]
    ws.append(["Field", "Value"])
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="4C3A20")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 118
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def build_outputs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    scenario_rows = _scenario_rows()
    decomposition_rows = _decomposition_rows()
    _write_csv(CSV_SCENARIOS, scenario_rows)
    _write_csv(CSV_DECOMPOSITION, decomposition_rows)

    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, "S6A_Top_Synergies", scenario_rows)
    _add_sheet(wb, "S6B_Decomposition", decomposition_rows)
    _add_metadata_sheet(wb)
    wb.save(XLSX_PATH)


if __name__ == "__main__":
    build_outputs()
    print(XLSX_PATH)
    print(CSV_SCENARIOS)
    print(CSV_DECOMPOSITION)
