"""Generate manuscript Table 3 and supplementary flux-ratio audit tables."""

from __future__ import annotations

import csv
import sys
from dataclasses import asdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from ExposoGraph.flux_engine import compute_pathway_flux  # noqa: E402

OUT_DIR = ROOT / "Supplementary_Tables"
XLSX_PATH = OUT_DIR / "Supplementary_Table_S4_ExposoGraph_flux_ratios.xlsx"
CSV_TABLE3 = OUT_DIR / "Table_3_representative_ExposoGraph_flux_ratios.csv"
CSV_TERMS = OUT_DIR / "Supplementary_Table_S4_flux_terms_long.csv"


SCENARIOS: list[dict[str, Any]] = [
    {
        "carcinogen_class": "PAH",
        "display_class": "PAH (benzo[a]pyrene)",
        "substrate": "Benzo[a]pyrene",
        "tissue": "Liver",
        "reference_label": "Reference",
        "reference_genotypes": {
            "CYP1A1": "NM",
            "CYP1B1": "NM",
            "EPHX1": "NM",
            "GSTM1": "active",
            "GSTP1": "NM",
        },
        "high_label": "High-susceptibility supported scenario",
        "high_genotypes": {
            "CYP1A1": "*2A/*2A",
            "CYP1B1": "NM",
            "EPHX1": "NM",
            "GSTM1": "null",
            "GSTP1": "NM",
        },
        "primary_driver": "CYP1A1 activation with GSTM1-null attenuation of GSH conjugation",
        "interpretation_note": (
            "CYP1A1 *2C and GSTP1 *B strings are not recognized by the current API."
        ),
    },
    {
        "carcinogen_class": "Aflatoxin",
        "display_class": "Aflatoxin B1",
        "substrate": "Aflatoxin B1",
        "tissue": "Liver",
        "reference_label": "Reference",
        "reference_genotypes": {
            "CYP3A4": "NM",
            "CYP1A2": "NM",
            "GSTA1": "NM",
        },
        "high_label": "High-susceptibility model-stress scenario",
        "high_genotypes": {
            "CYP3A4": "UM",
            "CYP1A2": "UM",
            "GSTA1": "PM",
        },
        "primary_driver": "CYP3A4/CYP1A2 activation with estimated GSTA1 detoxification term",
        "interpretation_note": "Detoxification includes a low-confidence estimated GSTA1 term.",
    },
    {
        "carcinogen_class": "AromaticAmines",
        "display_class": "Aromatic amines (4-ABP proxy)",
        "substrate": "4-aminobiphenyl proxy",
        "tissue": "Liver",
        "reference_label": "Reference",
        "reference_genotypes": {
            "CYP1A2": "NM",
            "NAT1": "NM",
            "NAT2": "rapid",
            "GSTM1": "active",
            "GSTP1": "NM",
            "XPC": "NM",
            "ERCC2": "NM",
        },
        "high_label": "High-susceptibility supported scenario",
        "high_genotypes": {
            "CYP1A2": "UM",
            "NAT1": "UM",
            "NAT2": "slow",
            "GSTM1": "null",
            "GSTP1": "PM",
            "XPC": "PM",
            "ERCC2": "PM",
        },
        "primary_driver": (
            "Slow NAT2 plus increased CYP1A2/NAT1 activation and reduced GST/NER attenuation"
        ),
        "interpretation_note": "This is a semi-quantitative proxy model, not measured kinetics.",
    },
    {
        "carcinogen_class": "Aldehyde",
        "display_class": "Acetaldehyde",
        "substrate": "Acetaldehyde/ethanol",
        "tissue": "Liver",
        "reference_label": "Reference",
        "reference_genotypes": {
            "ADH1B": "*1/*1",
            "ALDH2": "*1/*1",
            "ALDH1A1": "NM",
        },
        "high_label": "High-susceptibility supported scenario",
        "high_genotypes": {
            "ADH1B": "*2/*2",
            "ALDH2": "*2/*2",
            "ALDH1A1": "PM",
        },
        "primary_driver": "ADH1B rapid ethanol oxidation with impaired ALDH2/ALDH1A1 clearance",
        "interpretation_note": (
            "ALDH2 *2 drives impaired clearance; ALDH1A1 PM removes backup clearance."
        ),
    },
]


ESTROGEN_TISSUE_SCENARIOS = [
    ("Liver", {}),
    ("Breast", {}),
]


def _format_genotypes(genotypes: dict[str, str]) -> str:
    return "; ".join(f"{gene}={value}" for gene, value in genotypes.items())


def _result_row(
    scenario: dict[str, Any],
    label: str,
    genotypes: dict[str, str],
) -> dict[str, Any]:
    result = compute_pathway_flux(
        scenario["carcinogen_class"],
        genotypes,
        tissue=scenario["tissue"],
    )
    return {
        "display_class": scenario["display_class"],
        "api_class": scenario["carcinogen_class"],
        "scenario": label,
        "substrate": scenario["substrate"],
        "tissue": scenario["tissue"],
        "genotypes": _format_genotypes(genotypes),
        "substrate_concentration_uM": result.substrate_concentration_uM,
        "total_activation": result.total_activation,
        "total_detox": result.total_detox,
        "net_ratio": result.net_ratio,
        "risk_classification": result.risk_classification.value,
        "model_kind": result.model_kind,
        "parameter_source": result.parameter_source,
        "tissue_weight_source": result.tissue_weight_source.value,
        "warnings": "; ".join(result.warnings),
        "unit_note": result.unit_note,
    }


def _main_rows() -> list[dict[str, Any]]:
    rows = []
    for scenario in SCENARIOS:
        ref = _result_row(
            scenario,
            scenario["reference_label"],
            scenario["reference_genotypes"],
        )
        high = _result_row(scenario, scenario["high_label"], scenario["high_genotypes"])
        fold = high["net_ratio"] / ref["net_ratio"] if ref["net_ratio"] else None
        rows.append(
            {
                "carcinogen_class": scenario["display_class"],
                "tissue": scenario["tissue"],
                "model_kind": ref["model_kind"],
                "reference_genotype": ref["genotypes"],
                "reference_flux_ratio": ref["net_ratio"],
                "high_susceptibility_genotype": high["genotypes"],
                "high_susceptibility_flux_ratio": high["net_ratio"],
                "fold_change": round(fold, 1) if fold is not None else None,
                "primary_driver": scenario["primary_driver"],
                "caveat": scenario["interpretation_note"],
            }
        )
    return rows


def _scenario_rows() -> list[dict[str, Any]]:
    rows = []
    for scenario in SCENARIOS:
        rows.append(
            _result_row(
                scenario,
                scenario["reference_label"],
                scenario["reference_genotypes"],
            )
        )
        rows.append(_result_row(scenario, scenario["high_label"], scenario["high_genotypes"]))

    for tissue, genotypes in ESTROGEN_TISSUE_SCENARIOS:
        result = compute_pathway_flux("EstrogenMetabolites", genotypes, tissue=tissue)
        rows.append(
            {
                "display_class": "Estrogen metabolites",
                "api_class": "EstrogenMetabolites",
                "scenario": f"Reference tissue contrast: {tissue}",
                "substrate": "Estradiol/metabolite proxy",
                "tissue": tissue,
                "genotypes": _format_genotypes(genotypes) if genotypes else "default NM",
                "substrate_concentration_uM": result.substrate_concentration_uM,
                "total_activation": result.total_activation,
                "total_detox": result.total_detox,
                "net_ratio": result.net_ratio,
                "risk_classification": result.risk_classification.value,
                "model_kind": result.model_kind,
                "parameter_source": result.parameter_source,
                "tissue_weight_source": result.tissue_weight_source.value,
                "warnings": "; ".join(result.warnings),
                "unit_note": result.unit_note,
            }
        )
    return rows


def _term_rows() -> list[dict[str, Any]]:
    rows = []
    for scenario in SCENARIOS:
        for label_key, genotype_key in [
            ("reference_label", "reference_genotypes"),
            ("high_label", "high_genotypes"),
        ]:
            result = compute_pathway_flux(
                scenario["carcinogen_class"],
                scenario[genotype_key],
                tissue=scenario["tissue"],
            )
            for term_type, terms in [
                ("activation", result.activation_enzymes),
                ("detox_or_repair", result.detox_enzymes),
            ]:
                for term in terms:
                    row = asdict(term)
                    rows.append(
                        {
                            "display_class": scenario["display_class"],
                            "api_class": scenario["carcinogen_class"],
                            "scenario": scenario[label_key],
                            "term_type": term_type,
                            **row,
                        }
                    )
    return rows


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return "; ".join(f"{key}={val}" for key, val in value.items())
    return value


def _add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([_cell_value(row.get(header)) for header in headers])

    header_fill = PatternFill("solid", fgColor="563D7C")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for column_cells in ws.columns:
        header = str(column_cells[0].value)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200]
        )
        width = min(max(max_len + 2, len(header) + 2, 10), 55)
        ws.column_dimensions[column_cells[0].column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName=f"T_{title[:24]}", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium5", showRowStripes=True)
    ws.add_table(table)


def _add_metadata_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Metadata_References")
    rows = [
        ("Generated", date.today().isoformat()),
        ("API", "ExposoGraph.flux_engine.compute_pathway_flux"),
        ("Default tissue-weight source", "curated"),
        ("Measured parameter source", "ExposoGraph/data/kinetic_parameters.json"),
        ("Proxy parameter source", "ExposoGraph/data/proxy_flux_parameters.json"),
        ("Measured provenance source", "ExposoGraph/data/parameter_provenance.json"),
        ("Proxy provenance source", "ExposoGraph/data/proxy_flux_provenance.json"),
        (
            "Caveat",
            "Ratios are model outputs for specified genotype scenarios and default substrate "
            "concentrations; proxy classes contain estimated terms.",
        ),
        (
            "Unsupported draft genotypes",
            "CYP1A1 *2C/*2C and GSTP1 *B/*B are not currently recognized by "
            "genotype_modifier(); use supported phenotype labels or add mappings first.",
        ),
        (
            "HCA caveat",
            "The HCA measured-kinetics model currently sets NAT2_acetylation as a fixed "
            "20% of activation, so NAT2 genotype does not alter the HCA net ratio. "
            "Use AromaticAmines for the current NAT2-sensitive proxy.",
        ),
    ]
    ws.append(["Field", "Value"])
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="563D7C")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def build_outputs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    main_rows = _main_rows()
    scenario_rows = _scenario_rows()
    term_rows = _term_rows()

    _write_csv(CSV_TABLE3, main_rows)
    _write_csv(CSV_TERMS, term_rows)

    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, "Table3_Main", main_rows)
    _add_sheet(wb, "S4A_Scenarios", scenario_rows)
    _add_sheet(wb, "S4B_Enzyme_Terms", term_rows)
    _add_metadata_sheet(wb)
    wb.save(XLSX_PATH)


if __name__ == "__main__":
    build_outputs()
    print(XLSX_PATH)
    print(CSV_TABLE3)
    print(CSV_TERMS)
