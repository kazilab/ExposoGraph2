"""Generate manuscript Table 4 and supplementary exposure-risk audit tables."""

from __future__ import annotations

import csv
import sys
from dataclasses import asdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from ExposoGraph.exposure_engine import (  # noqa: E402
    compute_exposure_weighted_risk,
    get_exposure_scenarios,
)

OUT_DIR = ROOT / "Supplementary_Tables"
XLSX_PATH = OUT_DIR / "Supplementary_Table_S5_ExposoGraph_exposure_scores.xlsx"
CSV_TABLE4 = OUT_DIR / "Table_4_exposure_dependent_risk_reclassification.csv"
CSV_SCENARIOS = OUT_DIR / "Supplementary_Table_S5_exposure_scenarios_audit.csv"


TABLE4_CASES: list[dict[str, Any]] = [
    {
        "genotype_label": "GSTM1-null",
        "carcinogen_class": "PAH",
        "genotypes": {"GSTM1": "null"},
        "tissue": "Lung",
        "scenario_id": "general_population",
        "manuscript_scenario": "Office worker, non-smoker",
    },
    {
        "genotype_label": "GSTM1-null",
        "carcinogen_class": "PAH",
        "genotypes": {"GSTM1": "null"},
        "tissue": "Lung",
        "scenario_id": "smoker",
        "manuscript_scenario": "Current smoker",
    },
    {
        "genotype_label": "GSTM1-null",
        "carcinogen_class": "PAH",
        "genotypes": {"GSTM1": "null"},
        "tissue": "Lung",
        "scenario_id": "smoker_heavy_grilled_meat",
        "manuscript_scenario": "Smoker + daily grilled meat",
    },
    {
        "genotype_label": "ALDH2*1/*2",
        "carcinogen_class": "Aldehyde",
        "genotypes": {"ALDH2": "*1/*2"},
        "tissue": "Liver",
        "scenario_id": "nondrinker",
        "manuscript_scenario": "Non-drinker",
    },
    {
        "genotype_label": "ALDH2*1/*2",
        "carcinogen_class": "Aldehyde",
        "genotypes": {"ALDH2": "*1/*2"},
        "tissue": "Liver",
        "scenario_id": "light_drinker",
        "manuscript_scenario": "Light drinker (1-7/wk)",
    },
    {
        "genotype_label": "ALDH2*1/*2",
        "carcinogen_class": "Aldehyde",
        "genotypes": {"ALDH2": "*1/*2"},
        "tissue": "Liver",
        "scenario_id": "moderate_drinker",
        "manuscript_scenario": "Moderate drinker (7-14/wk)",
    },
]


AUDIT_CASES: list[dict[str, Any]] = [
    *TABLE4_CASES,
    {
        "genotype_label": "Reference PAH",
        "carcinogen_class": "PAH",
        "genotypes": {},
        "tissue": "Lung",
        "scenario_id": "general_population",
        "manuscript_scenario": "Reference PAH baseline",
    },
    {
        "genotype_label": "CYP1A1-high + GSTM1-null",
        "carcinogen_class": "PAH",
        "genotypes": {"CYP1A1": "high", "GSTM1": "null"},
        "tissue": "Lung",
        "scenario_id": "smoker",
        "manuscript_scenario": "Current smoker, high PAH susceptibility",
    },
    {
        "genotype_label": "CYP2E1-rapid + GSTT1-null",
        "carcinogen_class": "Benzene",
        "genotypes": {"CYP2E1": "rapid", "GSTT1": "null"},
        "tissue": "Liver",
        "scenario_id": "occupational_petroleum",
        "manuscript_scenario": "Petroleum/chemical worker",
    },
]


def _format_genotypes(genotypes: dict[str, str]) -> str:
    if not genotypes:
        return "reference/default"
    return "; ".join(f"{gene}={value}" for gene, value in genotypes.items())


def _case_row(case: dict[str, Any]) -> dict[str, Any]:
    result = compute_exposure_weighted_risk(
        case["carcinogen_class"],
        case["genotypes"],
        case["tissue"],
        exposure_scenario=case["scenario_id"],
    )
    effective_multiplier = result.exposure_multiplier * result.tissue_factor
    return {
        "genotype": case["genotype_label"],
        "carcinogen_class": result.carcinogen_class,
        "tissue": result.tissue,
        "genotypes_used": _format_genotypes(case["genotypes"]),
        "exposure_scenario": case["manuscript_scenario"],
        "exposure_scenario_id": result.exposure_scenario,
        "scenario_label": result.scenario_label,
        "exposure_tier": result.exposure_tier,
        "exposure_multiplier": result.exposure_multiplier,
        "tissue_factor": result.tissue_factor,
        "effective_exposure_tissue_multiplier": round(effective_multiplier, 4),
        "genotype_flux_ratio": result.flux_ratio,
        "combined_score": result.combined_risk_score,
        "risk_category": result.risk_category.value,
        "tissue_conc_uM": result.tissue_conc_uM,
        "sources": "; ".join(source for source in result.sources if source),
    }


def _table4_rows() -> list[dict[str, Any]]:
    return [_case_row(case) for case in TABLE4_CASES]


def _audit_rows() -> list[dict[str, Any]]:
    return [_case_row(case) for case in AUDIT_CASES]


def _scenario_rows() -> list[dict[str, Any]]:
    rows = []
    for class_name in ["PAH", "Aldehyde", "Benzene"]:
        for scenario in get_exposure_scenarios(class_name):
            row = asdict(scenario)
            row["carcinogen_class"] = class_name
            rows.append(row)
    return rows


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    if isinstance(value, dict):
        return "; ".join(f"{key}={val}" for key, val in value.items())
    return value


def _add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([_cell_value(row.get(header)) for header in headers])

    header_fill = PatternFill("solid", fgColor="174A43")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for column_cells in ws.columns:
        header = str(column_cells[0].value)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200]
        )
        width = min(max(max_len + 2, len(header) + 2, 10), 58)
        ws.column_dimensions[column_cells[0].column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table = Table(displayName=f"T_{title[:24]}", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws.add_table(table)


def _add_metadata_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Metadata_References")
    rows = [
        ("Generated", date.today().isoformat()),
        ("API", "ExposoGraph.exposure_engine.compute_exposure_weighted_risk"),
        ("Formula", "combined_score = genotype_flux_ratio x exposure_multiplier x tissue_factor"),
        ("Exposure source", "ExposoGraph/data/exposure_database.json"),
        (
            "Risk thresholds",
            "Low <= 0.8; Population Average >0.8 to <=2.0; Elevated >2.0 to <=10.0; High >10.0",
        ),
        (
            "Important distinction",
            "The package reports scenario exposure_multiplier separately from tissue_factor. "
            "The effective multiplier equals exposure_multiplier x tissue_factor.",
        ),
        (
            "Benzene audit",
            "The current Benzene occupational_petroleum scenario has multiplier 150.0; "
            "CYP2E1 rapid + GSTT1-null in liver gives score 729.0, not 38.2.",
        ),
    ]
    ws.append(["Field", "Value"])
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="174A43")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 118
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def build_outputs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    table4_rows = _table4_rows()
    audit_rows = _audit_rows()
    scenario_rows = _scenario_rows()

    _write_csv(CSV_TABLE4, table4_rows)
    _write_csv(CSV_SCENARIOS, scenario_rows)

    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, "Table4_Main", table4_rows)
    _add_sheet(wb, "S5A_Audit_Cases", audit_rows)
    _add_sheet(wb, "S5B_Scenarios", scenario_rows)
    _add_metadata_sheet(wb)
    wb.save(XLSX_PATH)


if __name__ == "__main__":
    build_outputs()
    print(XLSX_PATH)
    print(CSV_TABLE4)
    print(CSV_SCENARIOS)
