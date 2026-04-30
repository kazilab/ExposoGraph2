"""Generate Supplementary Table S3 from ExposoGraph tissue-expression data."""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from ExposoGraph.reference_data import build_full_panel

ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "ExposoGraph" / "data" / "tissue_expression_data.json"
OUT_DIR = ROOT / "Supplementary_Tables"
XLSX_PATH = OUT_DIR / "Supplementary_Table_S3_ExposoGraph_tissue_expression.xlsx"
CSV_PANEL_LONG = OUT_DIR / "Supplementary_Table_S3_ExposoGraph_tier1_2_expression_long.csv"
CSV_TABLE2 = OUT_DIR / "Table_2_representative_ExposoGraph_tissue_weights.csv"

REPRESENTATIVE_GENES = [
    "CYP1A1",
    "CYP1A2",
    "CYP1B1",
    "CYP2E1",
    "CYP3A4",
    "GSTM1",
    "NAT2",
]
TABLE2_TISSUES = ["Liver", "Lung", "Prostate", "Bladder", "Colon", "Breast"]
MISSING_NOTE = "Not stored in current ExposoGraph tissue expression matrix."


def _load_data() -> dict[str, Any]:
    with DATA_PATH.open() as fh:
        return json.load(fh)


def _panel_nodes() -> list[Any]:
    return build_full_panel().nodes


def _node_metadata() -> dict[str, dict[str, Any]]:
    nodes = _panel_nodes()
    return {
        node.id: {
            "gene": node.id,
            "tier": node.tier,
            "phase_or_group": node.phase or node.group or "",
            "role": node.role or "",
            "package_tissue_annotation": node.tissue or "",
            "detail": node.detail or "",
        }
        for node in nodes
    }


def _source_tissue_label(meta: dict[str, Any], tissue: str) -> str:
    value = meta.get("source_tissue_map", {}).get(tissue, [])
    if isinstance(value, list):
        return "; ".join(value)
    return str(value)


def _panel_long_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    meta = data["metadata"]
    expression = data["expression"]
    weights = data["weights"]
    tissue_names = meta["tissues"]
    node_meta = _node_metadata()

    rows: list[dict[str, Any]] = []
    for gene, details in node_meta.items():
        available = gene in expression and gene in weights
        for tissue in tissue_names:
            rows.append(
                {
                    **details,
                    "tissue": tissue,
                    "source_gtex_tissue": _source_tissue_label(meta, tissue),
                    "nTPM": expression.get(gene, {}).get(tissue),
                    "exposograph_weight": weights.get(gene, {}).get(tissue),
                    "hpa_gtex_available_in_package": available,
                    "notes": "" if available else MISSING_NOTE,
                }
            )
    return rows


def _wide_rows(
    genes: list[str],
    tissues: list[str],
    data: dict[str, Any],
    value_key: str,
    include_panel_metadata: bool = True,
) -> list[dict[str, Any]]:
    matrix = data[value_key]
    node_meta = _node_metadata()
    rows: list[dict[str, Any]] = []

    for gene in genes:
        meta = node_meta.get(
            gene,
            {
                "gene": gene,
                "tier": "",
                "phase_or_group": "",
                "role": "",
                "package_tissue_annotation": "",
                "detail": "",
            },
        )
        row = {
            "gene": gene,
            "tier": meta["tier"] if include_panel_metadata else "",
            "phase_or_group": meta["phase_or_group"] if include_panel_metadata else "",
            "role": meta["role"] if include_panel_metadata else "",
            "package_tissue_annotation": (
                meta["package_tissue_annotation"] if include_panel_metadata else ""
            ),
            "hpa_gtex_available_in_package": gene in matrix,
        }
        for tissue in tissues:
            row[tissue] = matrix.get(gene, {}).get(tissue)
        row["notes"] = "" if gene in matrix else MISSING_NOTE
        rows.append(row)
    return rows


def _representative_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    expression = data["expression"]
    weights = data["weights"]
    node_meta = _node_metadata()
    rows = []
    for gene in REPRESENTATIVE_GENES:
        meta = node_meta[gene]
        row: dict[str, Any] = {
            "gene": gene,
            "tier": meta["tier"],
            "phase_or_group": meta["phase_or_group"],
            "role": meta["role"],
        }
        for tissue in TABLE2_TISSUES:
            row[f"{tissue}_weight"] = weights[gene][tissue]
        for tissue in TABLE2_TISSUES:
            row[f"{tissue}_nTPM"] = expression[gene][tissue]
        rows.append(row)
    return rows


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])

    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for column_cells in ws.columns:
        header = str(column_cells[0].value)
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200]
        )
        width = min(max(max_len + 2, len(header) + 2, 10), 42)
        ws.column_dimensions[column_cells[0].column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    table_ref = ws.dimensions
    table = Table(displayName=f"T_{title.replace('-', '_')[:24]}", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)


def _add_metadata_sheet(wb: Workbook, data: dict[str, Any]) -> None:
    meta = data["metadata"]
    ws = wb.create_sheet("Metadata_References")
    rows = [
        ("Generated", date.today().isoformat()),
        ("Generated from package file", str(DATA_PATH.relative_to(ROOT))),
        ("Package source", meta.get("source")),
        ("Source table", meta.get("source_table")),
        ("Source URL", meta.get("source_url")),
        ("Unit", meta.get("unit")),
        ("Source file date", meta.get("source_file_date")),
        ("Package data date", meta.get("date")),
        ("Weight normalization", meta.get("weight_normalization")),
        ("Weight threshold", meta.get("weight_threshold")),
        ("Weight note", meta.get("weight_note")),
        ("Genes in tissue-expression matrix", meta.get("genes_total")),
        ("Graph enzymes total in metadata", meta.get("graph_enzymes_total")),
        (
            "Unavailable graph enzymes listed in metadata",
            ", ".join(meta.get("hpa_unavailable_graph_enzymes", [])),
        ),
        ("Panel genes absent from current expression matrix", "GSTT1; SULT1E1"),
        ("HPA tissue data page", "https://www.proteinatlas.org/humanproteome/tissue/data"),
        (
            "HPA transcriptomics methods",
            "https://www.proteinatlas.org/humanproteome/tissue/method/transcriptomics",
        ),
        (
            "GTEx Consortium reference",
            "Science 2020;369(6509):1318-1330. doi:10.1126/science.aaz1776",
        ),
        (
            "Interpretation note",
            "nTPM and derived weights are transcript-expression proxies, not direct enzyme "
            "activity or catalytic capacity.",
        ),
        (
            "Interpretation note",
            "Weights equal zero when expression is below the package 1% max-expression "
            "cutoff; inspect nTPM before calling a gene absent.",
        ),
    ]
    ws.append(["Field", "Value"])
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E5F")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"


def build_outputs() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = _load_data()
    tissues = data["metadata"]["tissues"]
    panel_genes = [node.id for node in _panel_nodes()]
    all_genes = sorted(data["expression"])

    panel_long = _panel_long_rows(data)
    representative = _representative_rows(data)
    _write_csv(CSV_PANEL_LONG, panel_long)
    _write_csv(CSV_TABLE2, representative)

    wb = Workbook()
    wb.remove(wb.active)
    _add_sheet(wb, "Table2_Representative", representative)
    _add_sheet(wb, "S3A_Panel38_nTPM", _wide_rows(panel_genes, tissues, data, "expression"))
    _add_sheet(wb, "S3B_Panel38_Weights", _wide_rows(panel_genes, tissues, data, "weights"))
    _add_sheet(wb, "S3C_Panel38_Long", panel_long)
    _add_sheet(
        wb,
        "S3D_All59_nTPM",
        _wide_rows(all_genes, tissues, data, "expression", include_panel_metadata=False),
    )
    _add_sheet(
        wb,
        "S3E_All59_Weights",
        _wide_rows(all_genes, tissues, data, "weights", include_panel_metadata=False),
    )
    _add_metadata_sheet(wb, data)
    wb.save(XLSX_PATH)


if __name__ == "__main__":
    build_outputs()
    print(XLSX_PATH)
    print(CSV_PANEL_LONG)
    print(CSV_TABLE2)
